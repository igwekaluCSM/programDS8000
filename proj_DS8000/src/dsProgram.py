import sys
import os
import time as t
import openpyxl as op
import atexit
import socket
import telnetlib as tn

from socket import gethostbyname, gaierror
from socket import gethostbyname, timeout
from socket import error as socket_error
from openpyxl import load_workbook  
from PyQt5  import QtCore, QtGui, uic, QtWidgets
from asyncio.tasks import wait
from fileinput import filename
from distutils.command.upload import upload
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtCore import QThread, pyqtSignal

# initialization necessities
currDir = os.getcwd()
end = "\n"
charReturn = "\r\n"


class BPLPrinter:

    def __init__(self, excDat):
        self.excDat = excDat
    
    def changeBPLDirectory(self):
        for bplText in open(currDir + '\\BPL.xml', 'r'):
            if "insert_directory_here" in bplText:
                bplText = bplText.replace("insert_directory_here" , currDir + '/BPL.xsd')
                bplText = bplText.replace('\\', '/')
                print(bplText)


class BradyIPPrinter:
        
        def __init__(self, gui):
            self.gui = gui
                
        def startPrinter(self):
            self.gui.connectButton.setEnabled(True)
            
        def connectPrinter(self):
            self.gui.inputField.setEnabled(True)
            self.gui.consoleOutput.append("Input the Printer IP Address: " + end)
            self.lastOutput = "Input the Printer IP Address"
        
        def inputData(self,text):
            if(text == "Input the Printer IP Address"): 
                ipAddress = self.gui.inputField.text()
                self.gui.consoleOutput.append("Printer IP Address: " + ipAddress + end)
                self.connectCheck(ipAddress)
                self.gui.connectButton.setEnabled(False)
                self.gui.printAllButton.setEnabled(True)
                self.gui.printSingleButton.setEnabled(True)
                self.gui.disconnectButton.setEnabled(True)
            if(text == "Type the label you would like to Print: "):
                self.singleLabel = self.gui.inputField.text()
                self.printLabel(self.singleLabel) 
            if(text == "Are you sure you want to print all labels? (Y/N): "):
                self.allConfirm = self.gui.inputField.text()
                if ((self.allConfirm[0] == 'Y') or (self.allConfirm == 'y')):
                    try:
                        self.allLabelPrint()
                    except Exception as e:
                        self.gui.consoleOutput.append("GUI Crashed: %s\n" % e)
                else:
                    self.gui.consoleOutput.append("Cancelled Print..."+ end)
                    self.resetInput()
                    
        def resetInput(self):
            self.gui.inputField.setText("")
            self.gui.inputField.setEnabled(False)
            
        def connectCheck(self,IP):
            try:
                global printConnect
                self.printConnect= tn.Telnet(IP, port=9100, timeout=10)
                
            except gaierror:
                self.gui.consoleOutput.append("Incorrect IP Address....Try Again" + end)
                self.gui.printAllButton.setEnabled(False)
                self.gui.printSingleButton.setEnabled(False)
                self.connectPrinter()
            except timeout:
                self.gui.consoleOutput.append("Incorrect IP Address....Try Again" + end)
                self.gui.printAllButton.setEnabled(False)
                self.gui.printSingleButton.setEnabled(False)
                self.connectPrinter()
            except socket_error:
                self.gui.consoleOutput.append("Incorrect IP Address....Try Again" + end)
                self.gui.printAllButton.setEnabled(False)
                self.gui.printSingleButton.setEnabled(False)
                self.connectPrinter()
            else:
                self.gui.consoleOutput.append("Printer " +IP+ " Connected!" + end)
                self.gui.consoleOutput.append("To Continue, Print Label(s)" + end)
                self.resetInput()
        
        def singlePrint(self):
            self.gui.inputField.setEnabled(True)
            self.gui.consoleOutput.append("Type the label you would like to Print: " + end)
            self.lastOutput = "Type the label you would like to Print: "
            
        def allPrint(self):
            self.gui.inputField.setEnabled(True)
            self.gui.consoleOutput.append("Are you sure you want to print all labels? (Y/N): " + end)
            self.lastOutput = "Are you sure you want to print all labels? (Y/N): "
                     
        def printLabel(self,oneLabel):
            for line in open(currDir + '\\labelCode.txt', 'r'):
                if("insert label here" in line):
                    line = line.replace("insert label here", oneLabel)
                try:    
                    self.printConnect.write((line).encode("ascii"))
                except Exception as e:
                    self.gui.consoleOutput.append("GUI Crashed: %s\n" % e)
                else:
                    self.resetInput()
            self.gui.consoleOutput.append("Done Printing!" + end)
            
        def allLabelPrint(self):
            for line in open(currDir + '\\labelCollect.txt', 'r'):
                if("There are no Labels in" not in line):
                    tempLabel = line
                    for line in open(currDir + '\\labelCode.txt', 'r'):
                        if("insert label here" in line):
                            line = line.replace("insert label here", tempLabel)
                        try:    
                            self.printConnect.write((line).encode("ascii"))
                        except Exception as e:
                            self.gui.consoleOutput.append("GUI Crashed: %s\n" % e)
            self.gui.consoleOutput.append("Done Printing!" + end)
            self.resetInput()

class ExcelData:

    def __init__(self, excDat):
        self.excDat = excDat
    
    def createExcel(self):

        os.system(currDir + '\\excelDS8000.xlsm')
        self.openSheet()
    
    def openLabelFile(self):
        
        try:
            os.remove("labelCollect.txt")
            self.labelFile = open("labelCollect.txt", 'a')
        except WindowsError:
            self.labelFile = open("labelCollect.txt", 'a')
            
    def openSheet(self):
        self.newSheet = currDir + '\\excelDS8000.xlsm'
        self.excelSheet = op.Workbook()
        self.excelSheet = load_workbook(filename=self.newSheet)
        sheetList = self.excelSheet.sheetnames
        self.labelRange = self.excelSheet['Sheet1']
        self.openLabelFile()
        
                    
        for sheetNum in sheetList:
            self.labelRange = self.excelSheet[sheetNum]
            try:
                self.gatherLabels(self.labelRange,sheetNum)
            except Exception as e:
                print("GUI Crashed: %s\n" % e) 
        
        self.labelFile.close()
        
    def loadSheet(self, fileLocation):
        self.excelSheet = op.Workbook()
        self.excelSheet = load_workbook(filename=fileLocation)
        sheetList = self.excelSheet.sheetnames
        
        self.openLabelFile()
            
        for sheetNum in sheetList:
            self.labelRange = self.excelSheet[sheetNum]
            try:
                self.gatherLabels(self.labelRange,sheetNum)
            except Exception as e:
                print("GUI Crashed: %s\n" % e)
                          
        self.labelFile.close()
        
    def gatherLabels(self, labelImport,thisSheet):
        
        currSheet = thisSheet
        
        self.constLabelNum = "7"
        currColumn = "B"
        self.initialLabel = labelImport["B7"].value
        label = labelImport[currColumn + self.constLabelNum].value

        if (self.initialLabel == ""):
            self.labelFile.write("There are no Labels in " + currSheet + "!...Check Programming Sheet" + end)  
        else:
                
            self.iterateLabel(label, self.labelFile, currColumn, labelImport) 
            
        
            
    def iterateLabel(self, tempLabel, allLabels, column, sheet):
        tempLabNum = self.constLabelNum
        iterColumn = column
        currSheet = sheet 
        currFile = allLabels
        if iterColumn == "B":
            while(tempLabel != ""): 
                currFile.write(tempLabel +"\n")
                tempLabNum = int(tempLabNum)
                tempLabNum += 1
                tempLabNum = str(tempLabNum)
                currLabel = iterColumn + tempLabNum
                tempLabel = currSheet[currLabel].value
            iterColumn = "F"
            tempLabel = currSheet["F7"].value
            self.iterateLabel(tempLabel, currFile, iterColumn, currSheet)
        else:
            while(tempLabel != ""): 
                currFile.write(tempLabel +"\n")
                tempLabNum = int(tempLabNum)
                tempLabNum += 1
                tempLabNum = str(tempLabNum)
                currLabel = iterColumn + tempLabNum
                tempLabel = currSheet[currLabel].value
     
class NetworkAdapter:

    def __init__(self, gui):
        self.gui = gui 
        
    def setNetAdapter(self):
        
        self.gui.resetNetAdaptButton.setEnabled(True)
        self.gui.setIPAdaptButton.setEnabled(False)
        self.gui.skipNetButton.setEnabled(False)
        self.gui.newExcelButton.setEnabled(True)
        self.gui.uploadExcelButton.setEnabled(True)
        #os.system( currDir + "\\netSet.cmd")
        #os.system( currDir + "\\tryThis.cmd > netsetOutput.txt")
        #self.gui.fillOutput("\\netsetOutput.txt")
        #os.system( currDir + "\\disableWiFi.cmd > disWiFiOutput.txt")
        #self.gui.fillOutput("\\disWiFiOutput.txt")
        os.system(currDir + "\\ipSet.cmd > ipsetOutput.txt")
        self.gui.fillOutput("\\ipsetOutput.txt")
        '''
        os.system("FOR /F \"tokens=4\" %G IN ('netsh int show int ^|find \"Local\"') DO SET netName=%G")
        os.system("FOR /F \"tokens=4\" %G IN ('netsh int show int ^|find \"Ethernet\"') DO echo Result is [%G]")
        #os.system("SET netName=%G%")
        os.system("ECHO %G")
        os.system("netsh int set int ECHO %netName% enable")
        #os.system("netsh int set int \"Ethernet\" enable")
        '''
        
    def resetNetAdapter(self):
        self.gui.resetNetAdaptButton.setEnabled(False)
        self.gui.setIPAdaptButton.setEnabled(True)
        
        # os.system("netsh int ip set address 13 dhcp")
        os.system(currDir + "\\ipReset.cmd > resetOutput.txt")
        self.gui.fillOutput('\\resetOutput.txt')
        
    def skipNetSet(self):
        
        self.gui.resetNetAdaptButton.setEnabled(True)
        self.gui.setIPAdaptButton.setEnabled(False)
        self.gui.newExcelButton.setEnabled(True)
        self.gui.uploadExcelButton.setEnabled(True)
        self.gui.skipNetButton.setEnabled(False)
        
            
        
class GUIMainWindow(QtWidgets.QMainWindow):
    
    def __init__(self, app):
        super(GUIMainWindow, self).__init__()
        self.gui = self
        
        self.app = app
        uic.loadUi(currDir + '\DS8000.ui', self)
        self.setWindowTitle("DS8000 Programming")
        
        self.networkAdapt = NetworkAdapter(self)
        self.excDat = ExcelData(self)
        self.bradyPrint = BradyIPPrinter(self)
        self.gui.consoleOutput.append("Pressing \"Start\" will install necessary components for this tool.." + end)  
        
    '''        
        self.progressClass = ProgressThread(self)
        self.progressClass.statusChange.connect(self.updateLoad)
        
    def updateLoad(self,value):
        self.progressBar.setValue(value)
       
    '''
        
    def enableConfig(self):
        self.setIPAdaptButton.setEnabled(True)
        self.startButton.setEnabled(False)
        self.skipNetButton.setEnabled(True)
        self.gui.consoleOutput.append("Installing Necessary Python Components: " + end)
        os.system(currDir + "\\pythonInstall.cmd > installOutput.txt")
        self.fillOutput("\\installOutput.txt")
        

        
    def startConfig(self):
        self.networkAdapt.setNetAdapter()

    def resetConfig(self):
        self.networkAdapt.resetNetAdapter()
        self.fillOutput("\\resetNet.txt")

    def skipConfig(self):
        self.networkAdapt.skipNetSet()
        
    def newExcel(self):
        self.excDat.createExcel()
        self.uploadNewData()
        
    def fillOutput(self, fileName):
        outputFile = open(currDir + fileName)
        setLines = outputFile.readlines()
        outputFile.close()
        try:
            for line in setLines:
                if ("The filename, directory name, or volume label syntax is incorrect." in line):
                    self.networkAdapt.setNetAdapter()
                self.consoleOutput.append(line)
        except Exception as e:  
            print("GUI Crashed: %s\n" % e)
        
    def openUpload(self):
        try:
            self.upload = UploadWindow(self)
            self.upload.show()
            
        except Exception as e:  
            print("GUI Crashed: %s\n" % e)
            
    def showLabels(self):
        self.gui.consoleOutput.append("Labels Found: " + end)
        self.gui.fillOutput('\\labelCollect.txt')
        self.bradyPrint.startPrinter()
        
    def uploadNewData(self):
        self.excDat.loadSheet(self.excDat.newSheet)
        try:
            self.gui.showLabels()

        except Exception as e:
            print("GUI Crashed: %s\n" % e)
        
    def enterHit(self):
        if(self.gui.inputField.text() == ""):
            self.gui.consoleOutput.append("Answer Cannot Be Blank!...Try Again" + end)
        else: 
            self.bradyPrint.inputData(self.bradyPrint.lastOutput)    
        
    def connectBrady(self):
        self.bradyPrint.connectPrinter()

    def printSingle(self):
        self.bradyPrint.singlePrint()
        
    def printAll(self):
        self.bradyPrint.allPrint()
        
    def donePrinting(self):
        try:
            self.bradyPrint.printConnect.close()
            
        except Exception as e:  
            print("GUI Crashed: %s\n" % e)  
        self.bradyPrint.startPrinter()
        self.gui.printAllButton.setEnabled(False)
        self.gui.printSingleButton.setEnabled(False)
        self.gui.disconnectButton.setEnabled(False)
        self.bradyPrint.resetInput()
        self.gui.consoleOutput.append("Printer Disconnected..."+ end)

'''   
class ProgressThread(QtCore.QThread):
    
    statusChange = pyqtSignal(int)
        
    def run(self):
        loading = 0
        while loading < 100:
            loading += 0.00001
            self.statusChange.emit(loading)
'''
                    
class UploadWindow(QtWidgets.QDialog):
    
    def __init__(self, parent=GUIMainWindow):
        super(UploadWindow, self).__init__(parent)
        self.gui = self
        
        self.excDat = ExcelData(self)
        uic.loadUi(currDir + '\\uploadWindow.ui', self)
          
    def openBrowse(self):
        self.browseWindow = QtWidgets.QFileDialog
        self.browseDirectory = self.browseWindow.getOpenFileName(self, "", "", "*.xls *.xlsb *.xlsm *.xlsx *.csv")
        self.showDirectory()
        # print(self.browseDirectory[0])
        
    def showDirectory(self):
        try:
            self.directoryOutput.setText(self.browseDirectory[0])
        except Exception as e:  
            print("GUI Crashed: %s\n" % e)
        
    def uploadData(self):
        self.excDat.loadSheet(self.browseDirectory[0])
        try:
            self.parent().showLabels()

        except Exception as e:
            print("GUI Crashed: %s\n" % e)
        
        
def main():
    try:
        app = QtWidgets.QApplication(sys.argv)
        window = GUIMainWindow(app)
        window.show()
        sys.exit(app.exec())
    except Exception as e:
        print("GUI Crashed: %s\n" % e)
        
# end of class MyApp


if __name__ == "__main__":
    main()
